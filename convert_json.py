from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import json
teamname = input("Team: ")

wb = load_workbook(filename=f'./{teamname}_기록.xlsx')
ws = wb.active
ws2 = wb['외야수']
ws3 = wb['내야수']
ws4 = wb['투수']
hitter = []
pitcher = []

def hitter_data(sheet):
    sheet['A1'] = '순위'
    last_column = len(list(sheet.columns))
    last_row = len(list(sheet.rows))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[sheet[column_letter + str(1)].value] = sheet[column_letter + str(row)].value
        hitter.append(my_dict)
    

def pitcher_data(sheet):
    sheet['A1'] = '순위'
    last_column = len(list(sheet.columns))
    last_row = len(list(sheet.rows))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[sheet[column_letter + str(1)].value] = sheet[column_letter + str(row)].value
        pitcher.append(my_dict)


hitter_data(ws)
hitter_data(ws2)
hitter_data(ws3)
pitcher_data(ws4)

hitter_json = json.dumps(hitter, sort_keys=True, indent=4,ensure_ascii=False)
with open(f'./{teamname}_야수_기록.json', 'w', encoding='utf-8') as f:
    f.write(hitter_json)

pitcher_json = json.dumps(pitcher, sort_keys=True, indent=4,ensure_ascii=False)
with open(f'./{teamname}_투수_기록.json', 'w', encoding='utf-8') as f:
    f.write(pitcher_json)